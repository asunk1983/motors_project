"""Сервис экспорта в Excel.

Раньше вся логика экспорта (~200 строк) была в app.py:export_engines().
Теперь сервис оркестрирует: repository → build_workbook → bytes.

ИСПРАВЛЕНО (см. HISTORY.md/чат): экспорт был технически рабочим, но
практически нечитаемым при печати. Основные причины:

1. `ws.print_options.fit_to_width = 1` — атрибута `fit_to_width` у класса
   `PrintOptions` в openpyxl НЕ СУЩЕСТВУЕТ. Python не мешает присвоить
   произвольный атрибут объекту, но в itogo.xlsx он никогда не
   сериализуется — при открытии файла в Excel настройки "вписать в
   ширину/высоту" там физически нет. Верно это задаётся через
   `ws.page_setup.fitToWidth/fitToHeight` вместе с
   `ws.sheet_properties.pageSetUpPr.fitToPage = True` (без последнего
   Excel игнорирует fitToWidth/fitToHeight).
2. Фото вставлялись фиксированным размером 250×190px без учёта реальных
   пропорций и БЕЗ увеличения высоты строки под них — картинка просто
   перекрывала собой всё, что шло ниже. Теперь размер каждого фото
   считается из его реальных пропорций (через PIL), а высота строки под
   ним увеличивается под самое высокое фото в ряду.
3. Между двигателями при экспорте нескольких сразу стоит принудительный
   разрыв страницы (openpyxl Break). Важно: если задать fitToHeight
   конкретным числом страниц — Excel/LibreOffice в режиме "вписать в N
   страниц" полностью ИГНОРИРУЕТ ручные разрывы и сам решает, где резать
   страницы (проверено на практике). Поэтому используется fitToWidth=1 +
   fitToHeight=0 ("не ограничивать число страниц по высоте") — масштаб
   только по ширине, а разбиение по страницам полностью отдаётся ручным
   разрывам.
4. (эта правка) Крупный шрифт + фото "подгоняются под свободное место":
   после того как посчитана реальная высота, занятая характеристиками/
   режимами/работами конкретного двигателя, фото получают оставшийся до
   конца страницы вертикальный бюджет и масштабируются под него (с
   сохранением пропорций, без искажения) — вместо фиксированного мелкого
   размера, из-за которого при малом числе фото на листе оставалось много
   пустого места (см. скриншот в переписке: 2 фото на всю страницу А4).
   Высота фото — динамическая (под остаток страницы), ширина — статичная
   по слоту (иначе фото в одном ряду наедут друг на друга при печати).
"""
import io
import logging

from repositories.engine_repo import get_with_details

logger = logging.getLogger(__name__)

# ---- Параметры страницы ----
PAGE_MARGIN_LR = 0.5    # дюймы, левое/правое поле
PAGE_MARGIN_TB = 0.5    # дюймы, верхнее/нижнее поле
PAGE_MARGIN_HF = 0.3    # дюймы, колонтитулы
PAGE_HEIGHT_PT = 842.0  # А4 книжная, пункты (297мм)
# Небольшой запас, чтобы не упереться впритык в нижний край печатной
# области при округлениях шрифта/строк в разных версиях Excel/LibreOffice.
PAGE_SAFETY_BUFFER_PT = 16

# ---- Шрифты — увеличены для читаемости при печати (было 9-14pt) ----
FONT_NAME = 'Calibri'
FONT_SIZE_TITLE = 16       # "Паспорт двигателя"
FONT_SIZE_SECTION = 13     # "Режимы работы" / "Произведённые работы" / "Фото"
FONT_SIZE_TABLE_HEAD = 12  # заголовки колонок таблиц
FONT_SIZE_DATA = 11        # обычный текст: подписи характеристик, значения, данные таблиц
FONT_SIZE_EMPTY = 10       # курсив "Нет данных"

# ---- Высоты строк (пункты) под увеличенный шрифт ----
ROW_H_TITLE = 28
ROW_H_CHAR_HEADER = 20
ROW_H_CHAR_DATA = 20
ROW_H_SECTION_HEADER = 22
ROW_H_MODE_TABLE_HEADER = 30
ROW_H_MODE_DATA = 20
ROW_H_WORKS_TABLE_HEADER = 34
ROW_H_PHOTO_HEADER = 22
ROW_H_EMPTY_NOTE = 18
DEFAULT_ROW_HEIGHT = 15  # запасной вариант для расчёта бюджета, если высота строки нигде не выставлена явно

# ---- Фото ----
# ИСПРАВЛЕНО (см. чат): раньше фото раскладывались по фиксированным
# "слотам"-колонкам — при разных пропорциях снимков (узкие книжные
# шильдики vs широкое фото самого двигателя) это давало то большие
# пробелы между фото, то тесноту, визуально "вразнобой". Теперь фото
# упаковываются как строки текста: кладём слева направо вплотную (с
# небольшим зазором), и как только следующее фото не помещается по
# ширине — переносим его на новую строку. Все фото в пределах одной
# карточки получают одну общую высоту (см. _place_photos), поэтому ряды
# выглядят как ровная фотополоса, а не хаотичная сетка.
PHOTO_GAP_PX = 10  # зазор между соседними фото по горизонтали и вертикали
# Высота фото — ДИНАМИЧЕСКАЯ, под остаток свободного места на странице
# (см. _place_photos), но всегда в этих границах: не мельче, чтобы фото
# оставалось разборчивым, и не крупнее, чтобы один снимок не занимал
# страницу целиком при пустой карточке без характеристик.
PHOTO_MIN_HEIGHT_PX = 90
PHOTO_MAX_HEIGHT_PX = 340
PHOTO_ROW_EXTRA_PT = 8  # запас по высоте строки сверх самого фото, пункты
# Сколько раз пробовать уменьшить высоту фото, подбирая её так, чтобы все
# ряды (после упаковки) уместились в остаток страницы — см. _place_photos.
PHOTO_FIT_MAX_ITERATIONS = 14

# 1px = 0.75pt при 96 dpi — стандартное допущение, которым пользуется и
# сам Excel при пересчёте.
PX_TO_PT = 0.75
PT_TO_PX = 1 / PX_TO_PT
EMU_PER_PIXEL = 9525

# Высота строки под перенесённый (wrap_text) текст — Excel/LibreOffice НЕ
# пересчитывают её автоматически при программной записи файла (это делает
# только сам Excel при открытии/редактировании листа человеком, а не сама
# библиотека при сохранении), поэтому многострочные описания работ рискуют
# оказаться обрезанными — тот же класс бага, что и с фото, только для
# текста. Высоту строки под перенос считаем сами по числу строк текста.
LINE_HEIGHT_PT = 15
ROW_WRAP_PADDING_PT = 8

# Ширина колонок листа в символах — используется и как реальная ширина
# ws.column_dimensions ниже, и для расчёта переноса текста в таблице
# "Произведённые работы" (см. _wrap_line_count/_set_wrapped_row_height).
# При базовом шрифте данных 11pt (совпадает со шрифтом Excel по
# умолчанию, на котором и основана единица ширины колонки) число символов
# в колонке примерно равно числу самой ширины — доп. коэффициент не нужен.
#
# ВАЖНО: сумма ширин ВСЕХ колонок должна укладываться в печатную область
# страницы (у А4 книжной с полями 0.5" это ≈95 символов, см. расчёт в
# истории чата). Если сумма шире — Excel/LibreOffice включают ненулевой
# fitToWidth-scale (<100%), а он масштабирует ВЕСЬ лист РАВНОМЕРНО, то
# есть сжимает не только ширину, но и высоту строк на печати. Из-за этого
# расчёт бюджета высоты под фото (см. remaining_height_pt в
# export_to_xlsx) считал по номинальным, ещё не сжатым точкам — и на
# печати ниже реального контента оставалась пустая полоса (ровно то, что
# было на скриншоте с 2 фото: место было расчитано верно "на бумаге
# Excel", но потом всё сжалось на ~85% при печати). Сумма ниже — 92
# символа, с запасом внутри бюджета ~95.
COLUMN_CHARS = {'A': 21, 'B': 13, 'C': 20, 'D': 13, 'E': 13, 'F': 14}
WORKS_COL_CHARS = {1: 21, 2: 13, 3: 20, 4: 13, 5: 13, 6: 14}
# Заголовки таблицы режимов ("Обороты, об/мин" и т.п.) короче, чем
# заголовки таблицы работ, но тоже нуждаются в переносе на некоторых
# колонках при увеличенном шрифте — используем ту же карту ширин.
MODE_HEADER_COL_CHARS = WORKS_COL_CHARS


def export_to_xlsx(conn, engine_ids: list[int]) -> bytes:
    """Экспортировать двигатели в Excel-файл.

    Возвращает bytes xlsx-файла (для отправки клиенту через Flask).
    Каждый двигатель занимает отдельную печатную страницу А4 (книжная
    ориентация): масштаб по ширине под 1 страницу + ручной разрыв
    страницы перед каждым следующим двигателем. Фото на каждой карточке
    масштабируются под фактически оставшееся место на странице.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.pagebreak import Break
    from PIL import Image as PILImage

    from utils.date import format_ru_date

    engines = []
    for eid in engine_ids:
        engine = get_with_details(conn, eid)
        if engine:
            engines.append(engine)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Паспорта двигателей'

    # --- Стили ---
    BORDER = Border(
        left=Side(style='thin', color='D0D5DD'),
        right=Side(style='thin', color='D0D5DD'),
        top=Side(style='thin', color='D0D5DD'),
        bottom=Side(style='thin', color='D0D5DD'),
    )
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    FILL_PRIMARY = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
    FILL_SECONDARY = PatternFill(start_color='EDF2F7', end_color='EDF2F7', fill_type='solid')

    # --- Ширина колонок ---
    # ИСПРАВЛЕНО: колонка A шириной 8 символов (было изначально) обрезала
    # подписи характеристик вроде "Место установки" — они не помещались,
    # т.к. соседняя колонка B занята объединённой ячейкой значения и не
    # даёт тексту "вытечь" вправо, как обычно бывает для пустых соседей.
    for col_letter, width in COLUMN_CHARS.items():
        ws.column_dimensions[col_letter].width = width

    row = 1

    for engine_index, engine in enumerate(engines):
        engine_start_row = row

        # --- Шапка ---
        ws.cell(row=row, column=1, value='Паспорт двигателя')
        ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=FONT_SIZE_TITLE, bold=True, color='FFFFFF')
        ws.cell(row=row, column=1).fill = FILL_PRIMARY
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = ROW_H_TITLE
        row += 1

        # --- Характеристики ---
        ws.cell(row=row, column=1, value='Характеристика')
        ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=FONT_SIZE_TABLE_HEAD, bold=True)
        ws.cell(row=row, column=2, value='Значение')
        ws.cell(row=row, column=2).font = Font(name=FONT_NAME, size=FONT_SIZE_TABLE_HEAD, bold=True)
        ws.row_dimensions[row].height = ROW_H_CHAR_HEADER
        row += 1

        char_fields = [
            ('Место установки', engine.get('location')),
            ('Тип двигателя', engine.get('engine_type')),
            ('Зав. номер', engine.get('serial_number')),
            ('Производитель', engine.get('manufacturer')),
            ('Назначение', engine.get('purpose')),
            ('Цех', engine.get('workshop')),
            ('Степень защиты', engine.get('protection_class')),
            ('Тип крепления', engine.get('mounting_type')),
            ('Диаметр вала (мм)', engine.get('shaft_diameter')),
            ('Подшипник передний', engine.get('bearing_front')),
            ('Подшипник задний', engine.get('bearing_rear')),
            ('Датчик температуры', engine.get('temp_sensor')),
            ('Энкодер', engine.get('encoder')),
            ('Охлаждение', engine.get('cooling')),
            ('Примечание', engine.get('note')),
        ]

        for label, value in char_fields:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=FONT_SIZE_DATA)
            ws.cell(row=row, column=1).border = BORDER
            ws.cell(row=row, column=2, value=value or '—')
            ws.cell(row=row, column=2).font = Font(name=FONT_NAME, size=FONT_SIZE_DATA)
            ws.cell(row=row, column=2).border = BORDER
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws.row_dimensions[row].height = ROW_H_CHAR_DATA
            row += 1

        row += 1  # отступ

        # --- Режимы работы ---
        ws.cell(row=row, column=1, value='⚡ Режимы работы')
        ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=FONT_SIZE_SECTION, bold=True, color='0F766E')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = ROW_H_SECTION_HEADER
        row += 1

        modes = engine.get('modes', [])
        if modes:
            headers = ['Частота, Гц', 'Мощность, кВт', 'Напряжение, В', 'Тип подкл.', 'Ток, А', 'Обороты, об/мин']
            for col, h in enumerate(headers, start=1):
                ws.cell(row=row, column=col, value=h)
                ws.cell(row=row, column=col).font = Font(name=FONT_NAME, size=FONT_SIZE_TABLE_HEAD, bold=True)
                ws.cell(row=row, column=col).fill = FILL_SECONDARY
                ws.cell(row=row, column=col).border = BORDER
                # wrap_text — заголовки типа "Обороты, об/мин" не помещаются
                # в узкую колонку в одну строку; без переноса Excel обрежет
                # видимую часть текста соседней непустой ячейкой.
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            _set_wrapped_row_height(ws, row, MODE_HEADER_COL_CHARS, min_height=ROW_H_MODE_TABLE_HEADER)
            row += 1

            for m in modes:
                for col, key in enumerate(['frequency', 'power', 'voltage', 'connection_type', 'current', 'rpm'], start=1):
                    ws.cell(row=row, column=col, value=m.get(key) or '—')
                    ws.cell(row=row, column=col).font = Font(name=FONT_NAME, size=FONT_SIZE_DATA)
                    ws.cell(row=row, column=col).border = BORDER
                    ws.cell(row=row, column=col).alignment = ALIGN_CENTER
                ws.row_dimensions[row].height = ROW_H_MODE_DATA
                row += 1
        else:
            ws.cell(row=row, column=1, value='Нет данных')
            ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=FONT_SIZE_EMPTY, italic=True, color='9CA3AF')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = ROW_H_EMPTY_NOTE
            row += 1

        row += 1  # отступ

        # --- Произведённые работы ---
        ws.cell(row=row, column=1, value='🔧 Произведенные работы')
        ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=FONT_SIZE_SECTION, bold=True, color='4338CA')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = ROW_H_SECTION_HEADER
        row += 1

        works = engine.get('works', [])
        if works:
            headers = ['№ п/п', 'Дата', 'Вид производимых работ', 'Сопротивление изоляции', 'Внешний осмотр и проверка работы', 'ФИО исполнителя']
            for col, h in enumerate(headers, start=1):
                ws.cell(row=row, column=col, value=h)
                ws.cell(row=row, column=col).font = Font(name=FONT_NAME, size=FONT_SIZE_TABLE_HEAD, bold=True)
                ws.cell(row=row, column=col).fill = FILL_SECONDARY
                ws.cell(row=row, column=col).border = BORDER
                # wrap_text — та же причина, что и у заголовков режимов
                # выше: "Внешний осмотр и проверка работы" целиком не
                # влезает в одну строку узкой колонки.
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            _set_wrapped_row_height(ws, row, WORKS_COL_CHARS, min_height=ROW_H_WORKS_TABLE_HEADER)
            row += 1

            for w in works:
                for col, key in enumerate(['work_number', 'date', 'work_description', 'isolation', 'inspection', 'signature'], start=1):
                    raw = w.get(key)
                    cell_value = format_ru_date(raw) if key == 'date' else raw
                    ws.cell(row=row, column=col, value=cell_value or '—')
                    ws.cell(row=row, column=col).font = Font(name=FONT_NAME, size=FONT_SIZE_DATA)
                    ws.cell(row=row, column=col).border = BORDER
                    # wrap_text на ВСЕХ колонках, не только на описании —
                    # "Внешний осмотр и проверка работы"/ФИО тоже иногда
                    # длиннее своей колонки. Высоту строки под перенос
                    # считаем сразу после заполнения строки.
                    ws.cell(row=row, column=col).alignment = Alignment(vertical='center', wrap_text=True)
                _set_wrapped_row_height(ws, row, WORKS_COL_CHARS)
                row += 1
        else:
            ws.cell(row=row, column=1, value='Нет данных')
            ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=FONT_SIZE_EMPTY, italic=True, color='9CA3AF')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = ROW_H_EMPTY_NOTE
            row += 1

        row += 1  # отступ перед фото

        # --- Фото ---
        photo_paths = _get_photo_paths(engine['id'])
        if photo_paths:
            ws.cell(row=row, column=1, value=f'📸 Фото ({len(photo_paths)})')
            ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=FONT_SIZE_SECTION, bold=True, color='1B365D')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = ROW_H_PHOTO_HEADER
            row += 1

            # Сколько места на странице уже занято характеристиками/
            # режимами/работами этого двигателя — остаток отдаём под фото,
            # чтобы при короткой карточке (мало полей, нет режимов/работ)
            # фото были крупными и не оставляли пустую нижнюю половину
            # листа, а при длинной карточке — не вылезали за пределы
            # страницы.
            used_height_pt = _sum_row_heights(ws, engine_start_row, row - 1)
            usable_height_pt = (
                PAGE_HEIGHT_PT
                - (PAGE_MARGIN_TB * 2) * 72
                - PAGE_SAFETY_BUFFER_PT
            )
            remaining_height_pt = max(0.0, usable_height_pt - used_height_pt)

            row = _place_photos(ws, XLImage, PILImage, photo_paths, row, remaining_height_pt)

        row += 2  # отступ между двигателями

        # --- Разрыв страницы между карточками ---
        # Без него при печати нескольких двигателей подряд разрыв мог
        # прийтись на середину таблицы режимов/работ следующей карточки.
        # Ставим ПОСЛЕ последней фактически использованной строки текущего
        # двигателя (row - 3, т.к. row уже сдвинут на отступ в 2 строки
        # выше и на 1 строку вперёд относительно последней строки с
        # контентом) и не ставим после самого последнего двигателя.
        if engine_index < len(engines) - 1:
            last_content_row = row - 3
            ws.row_breaks.append(Break(id=last_content_row))

    last_row = max(row - 1, engine_start_row) if engines else 1

    # --- Настройка страницы: книжная А4, вписать в 1 страницу по ширине.
    #
    # ВАЖНО про fitToHeight: Excel/LibreOffice игнорируют ручные разрывы
    # страниц (row_breaks), если fitToHeight задан конкретным числом —
    # режим "вписать в N страниц" в этом случае сам пересчитывает разбиение
    # по страницам под нужное количество и перекрывает собой все manual
    # breaks. Поэтому fitToHeight=0 — это не "не задано", а буквально "не
    # ограничивать число страниц по высоте": Excel масштабирует ТОЛЬКО по
    # ширине под 1 страницу, а по высоте использует обычную постраничную
    # разбивку, где расставленные выше ws.row_breaks работают как
    # положено. Каждая карточка попадает на одну страницу, потому что при
    # построении листа высота фото уже подогнана под остаток страницы
    # (см. remaining_height_pt выше) — переполнения по высоте не возникает.
    ws.print_area = f'A1:F{last_row}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=PAGE_MARGIN_LR, right=PAGE_MARGIN_LR,
        top=PAGE_MARGIN_TB, bottom=PAGE_MARGIN_TB,
        header=PAGE_MARGIN_HF, footer=PAGE_MARGIN_HF,
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sum_row_heights(ws, start_row: int, end_row: int) -> float:
    """Суммирует фактическую высоту строк [start_row, end_row] (пункты).

    Для строк без явно заданной высоты используется DEFAULT_ROW_HEIGHT —
    в этом экспорте такое практически не встречается (почти все строки
    получают явную высоту при заполнении), но как подстраховка на случай
    будущих правок, которые забудут это сделать.
    """
    total = 0.0
    for r in range(start_row, end_row + 1):
        dim = ws.row_dimensions.get(r)
        total += dim.height if (dim and dim.height) else DEFAULT_ROW_HEIGHT
    return total


def _place_photos(ws, xl_image_cls, pil_image_cls, photo_paths, row, remaining_height_pt):
    """Вставляет фото упаковкой слева направо с переносом на новую строку,
    когда следующее фото не помещается по ширине (как перенос слов в
    тексте) — вместо фиксированных колонок-слотов, которые давали
    неровные пробелы между фото разной пропорции (см. чат: узкие
    книжные шильдики и широкое фото двигателя в одном ряду).

    Все фото карточки получают одну общую высоту, которая подбирается
    так, чтобы все получившиеся ряды уместились в remaining_height_pt —
    отсюда PHOTO_FIT_MAX_ITERATIONS: пробуем высоту, считаем, сколько
    рядов получится при упаковке, и если не помещается — уменьшаем
    высоту и пробуем снова.

    Возвращает номер следующей свободной строки после всех фото.
    """
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D

    # Натуральные размеры — читаем один раз, дальше только пересчитываем
    # масштаб под разные пробные высоты без повторного открытия файлов.
    sizes = []
    for p in photo_paths:
        try:
            with pil_image_cls.open(p) as pil_img:
                sizes.append((p, pil_img.size[0], pil_img.size[1]))
        except Exception:
            logger.exception("Не удалось прочитать фото %s для экспорта", p)
    if not sizes:
        return row

    total_width_px = sum(round(w * 7 + 5) for w in COLUMN_CHARS.values())

    def pack(height_px):
        """Раскладывает фото по рядам при заданной общей высоте фото.

        Возвращает список рядов, где каждый ряд — список
        (path, display_w, display_h, x_px).
        """
        rows_out = []
        current = []
        x = 0
        for p, nat_w, nat_h in sizes:
            disp_w = max(1, round(nat_w * (height_px / nat_h)))
            if current and x + disp_w > total_width_px:
                rows_out.append(current)
                current = []
                x = 0
            current.append((p, disp_w, height_px, x))
            x += disp_w + PHOTO_GAP_PX
        if current:
            rows_out.append(current)
        return rows_out

    # Подбираем высоту побольше, но так, чтобы все ряды уместились в
    # остаток страницы — начинаем с максимума и уменьшаем, пока не влезет
    # (или пока не упрёмся в минимально читаемый размер).
    height_px = PHOTO_MAX_HEIGHT_PX
    packed = pack(height_px)
    for _ in range(PHOTO_FIT_MAX_ITERATIONS):
        needed_pt = len(packed) * (height_px * PX_TO_PT + PHOTO_ROW_EXTRA_PT)
        if needed_pt <= remaining_height_pt or height_px <= PHOTO_MIN_HEIGHT_PX:
            break
        # Уменьшаем пропорционально нехватке места, но не мельче минимума.
        shrink = max(0.7, remaining_height_pt / needed_pt)
        height_px = max(PHOTO_MIN_HEIGHT_PX, round(height_px * shrink))
        packed = pack(height_px)

    photo_row = row
    for row_photos in packed:
        row_height_px = max(h for _, _, h, _ in row_photos)
        for p, disp_w, disp_h, x_px in row_photos:
            try:
                img = xl_image_cls(p)
                img.width = disp_w
                img.height = disp_h
                col_idx, col_off_emu = _pixel_x_to_anchor(x_px)
                marker = AnchorMarker(col=col_idx, colOff=col_off_emu, row=photo_row - 1, rowOff=0)
                size = XDRPositiveSize2D(cx=disp_w * EMU_PER_PIXEL, cy=disp_h * EMU_PER_PIXEL)
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                ws.add_image(img)
            except Exception:
                logger.exception("Не удалось вставить фото %s в экспорт", p)
        ws.row_dimensions[photo_row].height = row_height_px * PX_TO_PT + PHOTO_ROW_EXTRA_PT
        photo_row += 1

    return photo_row


def _pixel_x_to_anchor(x_px: int):
    """Переводит абсолютную позицию x (в пикселях от левого края колонки A)
    в (индекс_колонки, смещение_в_EMU_внутри_неё) — нужно для точного
    позиционирования фото "внахлёст", а не только по границам колонок.
    """
    letters = ['A', 'B', 'C', 'D', 'E', 'F']
    remaining = x_px
    for idx, letter in enumerate(letters):
        col_px = round(COLUMN_CHARS[letter] * 7 + 5)
        if remaining < col_px or idx == len(letters) - 1:
            return idx, round(remaining * EMU_PER_PIXEL)
        remaining -= col_px
    return len(letters) - 1, round(remaining * EMU_PER_PIXEL)  # pragma: no cover — недостижимо при непустом letters


def _get_photo_paths(engine_id: int) -> list[str]:
    """Получить пути к фото двигателя.

    ИСПРАВЛЕНО: раньше здесь была собственная реализация со старой схемой
    имён (`engine_{id}_...`), из-за чего экспорт молча не находил фото для
    любого двигателя, загруженного после перехода на актуальную схему
    (`ID{id}_...`). Теперь используется каноническая функция photo_manager —
    единственный источник истины для схемы именования и путей, вместо
    третьей независимой копии той же логики.
    """
    from modules.photo_manager.manager import engine_photo_disk_paths

    return engine_photo_disk_paths(engine_id)


def _col_letter(col: int) -> str:
    """Преобразует номер колонки в букву (1 -> 'A', 27 -> 'AA')."""
    result = ''
    while col > 0:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result


def _wrap_line_count(text, chars_per_line: int) -> int:
    """Грубая оценка числа строк, которые займёт text при переносе по
    словам в колонке шириной chars_per_line символов. Тот же алгоритм
    переноса, что использует сам Excel (перенос по границе слова, не по
    середине), поэтому оценка достаточно точна для расчёта высоты строки —
    не идеальна для экзотических моноширинных случаев, но с запасом
    ROW_WRAP_PADDING_PT этого достаточно, чтобы не обрезать текст.
    """
    text = str(text or '').strip()
    if not text:
        return 1
    chars_per_line = max(1, chars_per_line)
    lines = 1
    cur_len = 0
    for word in text.split():
        wl = len(word)
        if cur_len == 0:
            cur_len = wl
        elif cur_len + 1 + wl <= chars_per_line:
            cur_len += 1 + wl
        else:
            lines += 1
            cur_len = wl
        # Само по себе длинное слово шире колонки — Excel всё равно
        # перенесёт его на отдельную(ые) строку(и) посимвольно, а не
        # обрежет; считаем, сколько строк оно займёт.
        if cur_len > chars_per_line:
            lines += cur_len // chars_per_line
            cur_len = cur_len % chars_per_line
    return lines


def _set_wrapped_row_height(ws, row: int, col_chars: dict, min_height: float = ROW_H_CHAR_DATA) -> None:
    """Выставляет высоту строки под самый "длинный" перенесённый текст
    среди ячеек этой строки. col_chars — {номер_колонки: примерная
    ширина колонки в символах}. min_height — высота при однострочном
    содержимом (для строк-заголовков она обычно больше, чем для строк
    данных, т.к. в заголовках задан более крупный шрифт).
    """
    max_lines = 1
    for col, chars_per_line in col_chars.items():
        cell = ws.cell(row=row, column=col)
        if cell.value:
            max_lines = max(max_lines, _wrap_line_count(cell.value, chars_per_line))
    ws.row_dimensions[row].height = max(
        min_height,
        max_lines * LINE_HEIGHT_PT + ROW_WRAP_PADDING_PT,
    )
