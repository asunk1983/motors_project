# modules/engine_parser/parser.py
"""Парсинг Excel-паспортов электродвигателей.

Содержит функции для извлечения характеристик двигателя, режимов работы
и произведённых работ из .xlsx-файлов (через pandas/openpyxl), а также
обёртку parse_file_fast() для пакетного импорта.
"""
import os
import logging
import pandas as pd
import zipfile
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

from config.settings import PHOTOS_FOLDER, ALLOWED_PHOTO_EXT


def get_cell_safe(arr, r, c):
    try:
        if r >= arr.shape[0] or c >= arr.shape[1]:
            return ""
        val = arr[r, c]
        return str(val) if pd.notna(val) else ""
    except Exception:
        return ""


def get_cell_val_safe(arr, r, c):
    try:
        if r >= arr.shape[0] or c >= arr.shape[1]:
            return None
        val = arr[r, c]
        return str(val).strip() if pd.notna(val) else None
    except Exception:
        return None


def parse_engine_data(arr, filename):
    def get_cell(r, c):
        return get_cell_safe(arr, r, c)

    val_ceh = get_cell(10, 41)
    ceh_val = ""
    loc_val = ""
    if '№' in val_ceh:
        parts = val_ceh.split('№', 1)
        if len(parts) > 1:
            ceh_val = parts[1].split(' ', 1)[0] if parts[1].split(' ', 1) else ""
            loc_val = parts[1].split(' ', 1)[1] if len(parts[1].split(' ', 1)) > 1 else ""

    return {
        'filename': filename,
        'purpose': get_cell(9, 41),
        'workshop': ceh_val,
        'location': loc_val,
        'engine_type': get_cell(13, 50),
        'manufacturer': get_cell(14, 50),
        'serial_number': get_cell(15, 50),
        'bearing_front': get_cell(22, 50),
        'bearing_rear': get_cell(23, 50),
        'shaft_diameter': get_cell(24, 50),
        'protection_class': get_cell(25, 50),
        'mounting_type': get_cell(26, 50),
        'temp_sensor': get_cell(27, 50),
        'encoder': get_cell(28, 50),
        'cooling': get_cell(29, 50),
        'note': get_cell(30, 50)
    }


def parse_operating_modes(arr):
    def get_cell_val(r, c):
        return get_cell_val_safe(arr, r, c)

    modes = []
    start_col = 50
    for col in range(start_col, start_col + 20):
        freq = get_cell_val(16, col)
        if freq is None or freq == '':
            break
        modes.append({
            'frequency': freq,
            'power': get_cell_val(17, col) or '',
            'voltage': get_cell_val(18, col) or '',
            'connection_type': get_cell_val(19, col) or '',
            'current': get_cell_val(20, col) or '',
            'rpm': get_cell_val(21, col) or ''
        })
    return modes


def parse_maintenance_works(arr):
    def get_cell_val(r, c):
        return get_cell_val_safe(arr, r, c)

    works = []
    current_row_idx = 39  # Данные начинаются с 39 строки
    counter = 1  # Счетчик для нумерации работ, начиная с 1

    try:
        while current_row_idx < 500:
            # Проверяем наличие номера работы в колонке 13
            n_val = get_cell_val(current_row_idx, 13)

            # Если номер работы пустой - проверяем, не закончились ли данные
            if n_val is None or n_val == '':
                empty_count = 0
                for i in range(5):
                    check_row = current_row_idx + i
                    check_val = get_cell_val(check_row, 13)
                    if check_val is None or check_val == '':
                        empty_count += 1
                    else:
                        break
                if empty_count >= 5:
                    break
                current_row_idx += 1
                continue

            # === ПАРСИМ ДАТУ ===
            date_str = get_cell_val(current_row_idx, 15) or ""
            parsed_date = ""
            if date_str:
                # Пробуем разные форматы
                try:
                    # DD.MM.YY или DD.MM.YYYY
                    if '.' in date_str:
                        parts = date_str.split('.')
                        if len(parts) == 3:
                            day = parts[0].zfill(2)
                            month = parts[1].zfill(2)
                            year = parts[2]
                            if len(year) == 2:
                                year = '20' + year if int(year) < 50 else '19' + year
                            parsed_date = f"{year}-{month}-{day}"
                    # DD/MM/YY или DD/MM/YYYY
                    elif '/' in date_str:
                        parts = date_str.split('/')
                        if len(parts) == 3:
                            day = parts[0].zfill(2)
                            month = parts[1].zfill(2)
                            year = parts[2]
                            if len(year) == 2:
                                year = '20' + year if int(year) < 50 else '19' + year
                            parsed_date = f"{year}-{month}-{day}"
                    # Если не удалось распарсить - оставляем как есть
                    else:
                        parsed_date = date_str
                except:
                    parsed_date = date_str

            # === ПАРСИМ СОПРОТИВЛЕНИЕ ИЗОЛЯЦИИ ===
            isol_str = get_cell_val(current_row_idx, 41) or ""
            isol_value = ""
            if isol_str:
                # Убираем пробелы и заменяем запятую на точку
                clean = isol_str.replace(' ', '').replace(',', '.').replace('МОм', '').replace('мОм', '').strip()
                try:
                    # Пробуем преобразовать в число
                    float_val = float(clean)
                    isol_value = str(float_val)
                except:
                    # Если не число - оставляем как есть
                    isol_value = isol_str

            # === СОЗДАЕМ ЗАПИСЬ ===
            work = {
                'work_number': str(counter),  # СВОЯ нумерация, начиная с 1
                'date': parsed_date,
                'work_description': get_cell_val(current_row_idx, 19) or "",
                'isolation': isol_value,
                'inspection': get_cell_val(current_row_idx, 45) or "",
                'signature': get_cell_val(current_row_idx, 56) or ""
            }
            works.append(work)
            counter += 1

            current_row_idx += 1

    except Exception as e:
        print(f"  ⚠️ Ошибка парсинга работ: {e}")

    # Диагностика
    print(f"\n[ДИАГНОСТИКА] parse_maintenance_works вернула {len(works)} записей:")
    for i, w in enumerate(works):
        print(f"  {i+1}. work_number='{w.get('work_number', '')}', date='{w.get('date', '')}', isolation='{w.get('isolation', '')}', desc='{str(w.get('work_description', ''))[:30]}...'")

    return works


def extract_images_from_excel(file_path, filename, engine_id, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)

    try:
        image_count = 0

        try:
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                media_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/')]
                if media_files:
                    for idx, media_file in enumerate(media_files):
                        try:
                            ext = os.path.splitext(media_file)[1].lower()
                            if not ext or ext not in ALLOWED_PHOTO_EXT:
                                ext = '.png'
                            photo_filename = f"ID{engine_id}_{idx+1}{ext}"
                            photo_path = os.path.join(PHOTOS_FOLDER, photo_filename)
                            with zip_ref.open(media_file) as source:
                                with open(photo_path, 'wb') as target:
                                    target.write(source.read())
                            image_count += 1
                        except Exception as e:
                            logger.warning(f"Не удалось извлечь {media_file} из {filename}: {e}")
                    return image_count
        except zipfile.BadZipFile:
            pass

        try:
            wb = load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if hasattr(ws, '_images') and ws._images:
                    for idx, img in enumerate(ws._images):
                        try:
                            img_data = None
                            if hasattr(img, '_data'):
                                img_data = img._data()
                            elif hasattr(img, 'data'):
                                img_data = img.data
                            if img_data:
                                photo_filename = f"ID{engine_id}_{idx+1}.png"
                                photo_path = os.path.join(PHOTOS_FOLDER, photo_filename)
                                with open(photo_path, 'wb') as f:
                                    f.write(img_data)
                                image_count += 1
                        except Exception as e:
                            logger.warning(f"Не удалось сохранить изображение из {filename}: {e}")
            wb.close()
            return image_count
        except Exception as e:
            logger.warning(f"openpyxl не смог прочитать {filename}: {e}")

        return image_count
    except Exception as e:
        log(f"  ❌ Ошибка извлечения фото: {e}")
        return 0


def parse_file_fast(file_path, log_callback=None):
    def log(msg):
        if log_callback:
            log_callback(msg)

    try:
        filename = os.path.basename(file_path)
        if not os.path.exists(file_path):
            return {'success': False, 'filename': filename, 'error': 'Файл не найден'}

        try:
            df = pd.read_excel(file_path, sheet_name='Лист1', header=None)
        except:
            try:
                df = pd.read_excel(file_path, header=None)
            except Exception as e:
                return {'success': False, 'filename': filename, 'error': str(e)}

        if df.empty:
            return {'success': False, 'filename': filename, 'error': 'Файл пуст'}

        arr = df.to_numpy()
        engine_data = parse_engine_data(arr, filename)
        modes = parse_operating_modes(arr)
        works = parse_maintenance_works(arr)

        engine_tuple = (
            engine_data['filename'], engine_data['purpose'], engine_data['workshop'],
            engine_data['location'], engine_data['engine_type'], engine_data['manufacturer'],
            engine_data['serial_number'], engine_data['bearing_front'], engine_data['bearing_rear'],
            engine_data['shaft_diameter'], engine_data['protection_class'], engine_data['mounting_type'],
            engine_data['temp_sensor'], engine_data['encoder'], engine_data['cooling'],
            engine_data['note'], 0
        )

        return {
            'success': True,
            'filename': filename,
            'engine_tuple': engine_tuple,
            'modes': modes,
            'works': works,
            'file_path': file_path
        }
    except Exception as e:
        return {'success': False, 'filename': os.path.basename(file_path), 'error': str(e)}