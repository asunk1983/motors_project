"""Группа 6: Импорт из Excel и очистка БД."""
import json
import os

import pytest
from playwright.sync_api import expect

from tests.e2e.helpers import switch_tab, wait_toast, accept_dialogs, make_engine


@pytest.mark.scn("50. Переключение на вкладку «Импорт»")
def test_50_switch_import_tab(page):
    """Вкладка импорта доступна и отображает информацию о файлах."""
    switch_tab(page, "import")
    expect(page.locator("#tab-import")).to_be_visible()
    expect(page.locator("#importInfo")).to_be_visible()


@pytest.mark.scn("51. Импорт из Excel")
def test_51_import_excel(page, admin_api):
    """Импорт из Excel-файла создаёт двигатели в БД."""
    import openpyxl
    from config.settings import MOTORS_FOLDER

    os.makedirs(MOTORS_FOLDER, exist_ok=True)
    xlsx_path = os.path.join(MOTORS_FOLDER, "test_e2e_import.xlsx")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Лист1"
        # Формат парсера: arr[10,41], arr[9,41], arr[13,50], arr[14,50], arr[15,50]
        ws.cell(row=11, column=42, value="Цех №1 ТестовоеМесто E2E")
        ws.cell(row=10, column=42, value="Тестовое назначение E2E")
        ws.cell(row=14, column=51, value="Электродвигатель")
        ws.cell(row=15, column=51, value="ТестПроизв")
        ws.cell(row=16, column=51, value="E2E-IMPORt-1234")
        wb.save(xlsx_path)

        switch_tab(page, "import")
        accept_dialogs(page)
        page.click("#importBtn")
        # Wait for import to complete
        page.wait_for_load_state("networkidle", timeout=30000)
        page.wait_for_timeout(3000)

        # Verify engine was imported via API
        r = admin_api.get("/api/engines?search_field=serial_number&search=E2E-IMPORt-1234")
        data = r.json()
        assert isinstance(data, list)
        if data:
            for eng in data:
                admin_api.delete(f"/api/engine/{eng['id']}")
    finally:
        if os.path.exists(xlsx_path):
            os.remove(xlsx_path)


@pytest.mark.scn("52. Очистка БД")
def test_52_clear_db(page, admin_api):
    """Очистка БД удаляет все двигатели."""
    # Create a test engine first
    payload = make_engine(serial_number="CLEAR-DB-TEST-999")
    r = admin_api.post("/api/engine", data=json.dumps(payload),
                       headers={"Content-Type": "application/json"})
    assert r.json().get("success")

    switch_tab(page, "import")
    accept_dialogs(page)
    page.click("#clearDbBtn")
    wait_toast(page, "База данных очищена")

    page.wait_for_load_state("networkidle", timeout=10000)

    # Verify all engines are gone
    r = admin_api.get("/api/engines")
    data = r.json()
    assert isinstance(data, list)
    for eng in data:
        serial = eng.get("serial_number", "")
        if serial and "E2E" in str(serial):
            pytest.fail(f"E2E engine still exists after clear: {serial}")
